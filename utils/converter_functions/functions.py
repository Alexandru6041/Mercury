import datetime
import json
from pathlib import Path
from django import forms
import openpyxl
from string import Template
# from _math import add_1
from utils.converter_functions._math import add_1
from utils.exceptions import WrongTypeFieldException

class MappedColumn():
    '''
    Tip special pentru a deosebi o constanta de o coloana din doc .xlsx
    '''
    def __init__(self, col:str):
        self.col = col


def reformat(data:datetime.datetime):
    return data.strftime('%d.%m.%Y')

def load(sheet, RANGE:tuple, mapping:dict) -> dict:
    '''
    Deschide documentul xlsx si returneaza facturile sub forma de dict
    '''
    facturi = {}

    with open('templates/facturi/linie_template.xml', 'r') as f:
        t_line = Template(f.read())

    for i in range(RANGE[0], RANGE[1]):
        nrdoc = sheet[f"{mapping['nrdoc'].col}{i}"].value

        if not nrdoc in facturi:
            facturi[nrdoc] = {'line': i, 'facturi': []}
        
        factura = ''
        data = {
            'id': len(facturi[nrdoc]['facturi']) + 1,
            'pv': ''
        }

        for key in mapping:
            if type(mapping[key]) == MappedColumn:
                data[key] = sheet[f'{mapping[key].col}{i}'].value
            
            else: data[key] = mapping[key]

        try:
            factura = t_line.substitute(data)

        except TypeError as e:
            print(i)
            print(e)

        facturi[nrdoc]['facturi'].append(factura)

    for nrdoc in facturi:
        final = ''
        for factura in facturi[nrdoc]['facturi']:
            final += factura
            final += '\n'
        
        facturi[nrdoc]['final'] = final
    
    return facturi

def check_n_correct(xml:str) -> str:
    '''
    Verifica si modifica un str pentru a nu avea caractere ce nu apar in alfabetul englezesc
    '''
    x = 0

    while x < len(xml):
        if xml[x] == '&':
            xml = xml[:x] + ' ' + xml[x+1:]
        
        elif ord(xml[x].lower()) > 122:
            xml = xml[:x] + xml[x+1:]
            x -= 1

        x += 1
    
    return xml

def exists_sheet(file, sheet:str) -> bool:
    workbook = openpyxl.load_workbook(file, data_only=True)
    return (sheet in workbook.sheetnames)

def gen_xml(wb:str, sh:str, RANGE:tuple, mapping:dict, output_path:str) -> None:
    '''
    Genereaza fisierul .xml rezultat in urma procesarii datelor furnizate de client
    '''
    sheet = openpyxl.load_workbook(wb, data_only=True)[sh]
    facturi = load(sheet, RANGE, mapping)

    xml = ''

    with open('templates/facturi/template_factura.xml', 'r') as f:
        t_fact = Template(f.read())
    
    if type(mapping['nume_cli']) == str: # intrari
        tip = 0
        client = {'nume': mapping['nume_cli'], 'cif': mapping['cif_cli']}
    
    else: 
        tip = 1
        furnizor = {'nume': mapping['nume_fur'], 'cif': mapping['cif_fur']}

    for nrdoc in facturi:
        line = facturi[nrdoc]['line']

        if tip:
            client = {'nume': sheet[f'{mapping["nume_cli"].col}{line}'].value, 'cif': sheet[f'{mapping["cif_cli"].col}{line}'].value}

        else:
            furnizor = {'nume': sheet[f'{mapping["nume_fur"].col}{line}'].value, 'cif': sheet[f'{mapping["cif_fur"].col}{line}'].value}

        try:
            if type(mapping['data']) == MappedColumn:
                data = reformat(sheet[f'{mapping["data"].col}{line}'].value)
            
            else: data = mapping["data"]

        except AttributeError:
            raise WrongTypeFieldException('data')

        fact_final = t_fact.substitute({
            'nume_furnizor': furnizor['nume'],
            'cif_furnizor': furnizor['cif'],
            'nrreg_furnizor': '',
            'jud_furnizor': '',
            'adresa_furnizor': '',
            'nume_cli': client['nume'],
            'cif_cli': client['cif'],
            'nrreg_cli': "",
            'jud_cli': "",
            'adr_cli': "",
            'nr_doc': nrdoc,
            'data_doc': data,
            'tax_inv': '',
            'tva_incasare': '',
            'continut': facturi[nrdoc]['final']
        })
        xml += fact_final
        xml += '\n'
    
    with open('templates/facturi/base_template.xml', 'r') as f:
        base = Template(f.read())
    
    with open(output_path, 'w') as f:
        f.write(check_n_correct(base.substitute({'facturi': xml})))

def save_uploaded_xlsx(f, user_id:int):
    timestamp = datetime.datetime.timestamp(datetime.datetime.now())
    path = f'input files/{user_id}_{timestamp}.xlsx'

    with open(path, 'wb+') as d:
        for chunk in f.chunks():
            d.write(chunk)
    
    return path

def as_json(path:str, sh:str, header_row:int):
    '''
    Incarca fisierul .xlsx in format json pentru a putea fi afisat in pagina web
    '''
    sheet = openpyxl.load_workbook(path, data_only=True)[sh]
    header_col = 'A'

    while not sheet[f'{header_col}{header_row}'].value:
        header_col = add_1(header_col)
    
    headers = []
    first_col = header_col

    map = {
        'cols': [],
    }

    while sheet[f'{header_col}{header_row}'].value:
        map['cols'].append(header_col)
        headers.append(sheet[f'{header_col}{header_row}'].value)
        header_col = add_1(header_col)
    
    r = 1
    rows = []

    while r <= 10:
        col = first_col
        h = 0
        row = [header_row + r]

        while h < len(headers):
            value = sheet[f'{col}{header_row + r}'].value

            if type(value) == int and not value:
                pass

            elif not value:
                value = ''

            row.append(str(value))
            h += 1
            col = add_1(col)
        
        rows.append(row)
        r += 1

    return json.dumps({'headers': headers, 'rows': rows, 'map': map}).replace('\\n', ' ')

def get_output_path(id, cif_firma):
    name = f'F_RO{cif_firma}_multiple_{datetime.datetime.today().strftime("%d.%m.%Y")}.xml'
    output_dir = f'output files/{id}'
    output_path = f'{output_dir}/{name}'

    if not Path(output_dir).exists():
        Path(output_dir).mkdir()
    
    return output_path

def proccess(map:forms.Form, model, user_id):
    '''
    Genereaza fisierul xml si returneaza adresa lui
    '''
    path = f'input files/{model.nume}'
    data, l1, l2 = map.proccess(model.nume_firma, model.cif_firma) 
    output_path = get_output_path(user_id, model.cif_firma)
    
    gen_xml(path, model.sheet, (l1, l2+1), data, output_path)

    return output_path

if __name__ == '__main__':
    print(reformat('ceva'))
